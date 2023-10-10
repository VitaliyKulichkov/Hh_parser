import requests
import json
import re
import openpyxl


class Parser:

    def __init__(self):
        self.base_url = 'https://api.hh.ru/vacancies'
        self.pages = ''
        self.per_page = ''

    def get_number_of_pages(self, query: str) -> str:
        req = requests.get(self.base_url + f"?text={query}&area=1")
        json_text = json.loads(req.text)
        self.pages = json_text['pages']
        self.per_page = json_text['per_page']
        return self.pages, self.per_page

    def parse_by_name_of_vacancy(self, query: str):
        vacancy_info = []
        self.get_number_of_pages(query=query)
        for page in range(int(self.pages)-1):
            print(f'page: {page}')
            req = requests.get(self.base_url + f"?text={query}&page={page}&area=1")
            json_text = json.loads(req.text)
            print(f"text at page 0: {json_text['items']}")
            try:
                for i in range(int(self.per_page)):
                    name = json_text['items'][i]['name']
                    area = json_text['items'][i]['area']['name']
                    salary = json_text['items'][i]['salary']
                    experience = json_text['items'][i]['experience']['name']
                    schedule = json_text['items'][i]['schedule']['name']
                    description = json_text['items'][i]['snippet']['requirement']
                    employer = json_text['items'][i]['employer']['name']
                    if salary:
                        vacancy_info.append([name, area, salary, experience, schedule, description, employer])
            except KeyError:
                pass
        return vacancy_info

    def insert_to_excel(self, query: str):
        lst_vacancy_info = self.parse_by_name_of_vacancy(query)
        wb = openpyxl.Workbook()
        del wb['Sheet']
        new_sheet = wb.create_sheet('Вакансии')
        new_sheet.title = 'Вакансии'
        sheet = wb['Вакансии']
        headers = ['Название', 'Город', 'ЗП мин', 'ЗП макс', 'Опыт', 'Тип занятости', 'Требования']

        # columns dimensions
        sheet.column_dimensions["A"].width = 50
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 30
        sheet.column_dimensions["D"].width = 30
        sheet.column_dimensions["E"].width = 30
        sheet.column_dimensions["F"].width = 30
        sheet.column_dimensions["G"].width = 400

        for pos, val in enumerate(headers):
            sheet.cell(row=1, column=pos + 1).value = val

        for vac in range(len(lst_vacancy_info)):
            for i in range(len(lst_vacancy_info[vac])):
                # sheet.cell(row=i+1, column=i+1).value = vac[i]
                sheet.cell(row=vac + 1, column=1).value = lst_vacancy_info[vac][0]  # name
                sheet.cell(row=vac + 1, column=2).value = lst_vacancy_info[vac][1]  # area

                # min salary block
                if lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['from'] is not None:
                    sheet.cell(vac + 1, column=3).value = str(lst_vacancy_info[vac][2]['from']) + ' ' + str(
                        lst_vacancy_info[vac][2]['currency']) + ' ' + 'до вычета НДФЛ'
                elif lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['from'] is None:
                    sheet.cell(vac + 1, column=3).value = '-'
                elif not lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['from'] is not None:
                    sheet.cell(vac + 1, column=3).value = str(lst_vacancy_info[vac][2]['from']) + ' ' + str(
                        lst_vacancy_info[vac][2]['currency']) + ' ' + 'на руки'
                elif not lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['from'] is None:
                    sheet.cell(vac + 1, column=3).value = '-'

                # max salary block
                if lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['to'] is not None:
                    sheet.cell(vac + 1, column=4).value = str(lst_vacancy_info[vac][2]['to']) + ' ' + str(
                        lst_vacancy_info[vac][2]['currency']) + ' ' + 'до вычета НДФЛ'
                elif lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['to'] is None:
                    sheet.cell(vac + 1, column=4).value = '-'
                elif not lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['to'] is not None:
                    sheet.cell(vac + 1, column=4).value = str(lst_vacancy_info[vac][2]['to']) + ' ' + str(
                        lst_vacancy_info[vac][2]['currency']) + ' ' + 'на руки'
                elif not lst_vacancy_info[vac][2]['gross'] and lst_vacancy_info[vac][2]['to'] is None:
                    sheet.cell(vac + 1, column=4).value = '-'

                sheet.cell(row=vac + 1, column=5).value = lst_vacancy_info[vac][3]  # experience
                sheet.cell(row=vac + 1, column=6).value = lst_vacancy_info[vac][4]  # schedule
                sheet.cell(row=vac + 1, column=7).value = lst_vacancy_info[vac][5]  # requirements

        wb.save(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")


if __name__ == '__main__':
    # Parser().parse_by_name_of_vacancy('junior python developer')
    pass

