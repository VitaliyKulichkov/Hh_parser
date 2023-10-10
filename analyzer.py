from parser import Parser
from matplotlib import pyplot as plt
import numpy as np
import openpyxl


class Analyzer:

    def __init__(self, lst):
        # self.query = query
        # self.lst = Parser().parse_by_name_of_vacancy(query)
        self.lst_with_zero_exp = []
        self.lst_with_one_three_year_exp = []
        self.lst_with_three_six_year_exp = []
        self.lst_with_more_six_year_exp = []
        self.lst = lst

    def insert_into_excel(self, query: str):
        wb = openpyxl.load_workbook(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")
        if 'Анализ' not in wb.sheetnames:
            wb.create_sheet('Анализ')
            wb.save(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")

    def analyze(self):
        # lst = Parser().parse_by_name_of_vacancy(query)
        for vac in range(len(self.lst)):
            if self.lst[vac][3] == 'Нет опыта':
                self.lst_with_zero_exp.append(self.lst[vac])
            elif self.lst[vac][3] == 'От 1 года до 3 лет':
                self.lst_with_one_three_year_exp.append(self.lst[vac])
            elif self.lst[vac][3] == 'От 3 до 6 лет':
                self.lst_with_three_six_year_exp.append(self.lst[vac])
            else:
                self.lst_with_more_six_year_exp.append(self.lst[vac])

        return self.lst_with_zero_exp, self.lst_with_one_three_year_exp, self.lst_with_three_six_year_exp, self.lst_with_more_six_year_exp

    def calculate_average_salary(self):
        lst_with_min_salaries = []
        lst_with_max_salaries = []
        lst_with_average_salaries = []
        for i in range(len(self.analyze())):
            lst_with_min_salary_for_group = []
            lst_with_max_salary_for_group = []
            lst_with_zero_exp_for_group = self.analyze()[i]
            print(lst_with_zero_exp_for_group)
            if lst_with_zero_exp_for_group:
                for vac in range(len(lst_with_zero_exp_for_group)):
                    if (lst_with_zero_exp_for_group[vac][2]['from'] is not None and
                            lst_with_zero_exp_for_group[vac][2]['currency'] == 'RUR'):
                        lst_with_min_salary_for_group.append(lst_with_zero_exp_for_group[vac][2]['from'])
                    if (lst_with_zero_exp_for_group[vac][2]['to'] is not None and
                            lst_with_zero_exp_for_group[vac][2]['currency'] == 'RUR'):
                        lst_with_max_salary_for_group.append(lst_with_zero_exp_for_group[vac][2]['to'])

                if lst_with_min_salary_for_group and lst_with_max_salary_for_group:
                    average_min_salary = (np.mean(lst_with_min_salary_for_group)).round(0)
                    average_max_salary = (np.mean(lst_with_max_salary_for_group)).round(0)

                    average_salary = (average_max_salary + average_min_salary) / 2

                    lst_with_average_salaries.append(average_salary)
                    lst_with_min_salaries.append(average_min_salary)
                    lst_with_max_salaries.append(average_max_salary)

                elif not lst_with_min_salary_for_group and lst_with_max_salary_for_group:
                    average_max_salary = (np.mean(lst_with_max_salary_for_group)).round(0)
                    average_min_salary = 0
                    average_salary = (average_max_salary + average_min_salary) / 2
                    lst_with_average_salaries.append(average_salary)
                    lst_with_min_salaries.append(average_min_salary)
                    lst_with_max_salaries.append(average_max_salary)
                elif not lst_with_min_salary_for_group and not lst_with_max_salary_for_group:
                    average_min_salary = 0
                    average_max_salary = 0
                    average_salary = 0
                    lst_with_average_salaries.append(average_salary)
                    lst_with_min_salaries.append(average_min_salary)
                    lst_with_max_salaries.append(average_max_salary)

                # print(max_salary, min_salary, average_salary)
            else:
                lst_with_average_salaries.append(0)
                lst_with_max_salaries.append(0)
                lst_with_min_salaries.append(0)
        return lst_with_max_salaries, lst_with_min_salaries, lst_with_average_salaries

    def draw_max_salary_histogram(self, query: str):
        lst_with_max_salaries = self.calculate_average_salary()[0]
        plt.figure(figsize=(7, 5))
        x = ['Нет опыта', '1-3 года', '3-6 лет', 'от 6 лет']
        y = lst_with_max_salaries
        plt.scatter(x, y, marker=".", s=100, edgecolors="black", c="yellow")
        plt.title('Максимальная ЗП')
        plt.xlabel('Опыт работы, лет', fontweight='bold', color='black',
                   fontsize='12', horizontalalignment='center')
        plt.ylabel('Зарплата, рублей', fontweight='bold', color='black', fontsize='12', horizontalalignment='center')
        plt.plot(x, y, '-o')
        # plt.show()
        plt.savefig(f'Максимальная ЗП.png')
        img = openpyxl.drawing.image.Image(f'Максимальная ЗП.png')
        self.insert_into_excel(query)
        wb = openpyxl.load_workbook(f"{query.replace(' ', '_')}_vacancies_HH.xlsx")
        sheet = wb['Анализ']
        img = openpyxl.drawing.image.Image(f'Максимальная ЗП.png')
        img.anchor = 'A1'
        sheet.add_image(img)
        wb.save(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")

    def draw_min_salary_histogram(self, query: str):
        lst_with_max_salaries = self.calculate_average_salary()[1]
        plt.figure(figsize=(7, 5))
        x = ['Нет опыта', '1-3 года', '3-6 лет', 'от 6 лет']
        y = lst_with_max_salaries
        plt.scatter(x, y, marker=".", s=100, edgecolors="black", c="yellow")
        plt.title('Минимальная ЗП')
        plt.xlabel('Опыт работы, лет', fontweight='bold', color='black',
                   fontsize='12', horizontalalignment='center')
        plt.ylabel('Зарплата, рублей', fontweight='bold', color='black', fontsize='12', horizontalalignment='center')
        plt.plot(x, y, '-o')
        # plt.show()
        plt.savefig(f'Минимальная ЗП.png')
        img = openpyxl.drawing.image.Image(f'Минимальная ЗП.png')
        self.insert_into_excel(query)
        wb = openpyxl.load_workbook(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")
        sheet = wb['Анализ']
        img = openpyxl.drawing.image.Image(f'Минимальная ЗП.png')
        img.anchor = 'C1'
        sheet.add_image(img)
        wb.save(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")

    def draw_average_salary_histogram(self, query: str):
        lst_with_max_salaries = self.calculate_average_salary()[2]
        plt.figure(figsize=(7, 5))
        x = ['Нет опыта', '1-3 года', '3-6 лет', 'от 6 лет']
        y = lst_with_max_salaries
        plt.scatter(x, y, marker=".", s=100, edgecolors="black", c="yellow")
        plt.title('Средняя ЗП')
        plt.xlabel('Опыт работы, лет', fontweight='bold', color='black',
                   fontsize='12', horizontalalignment='center')
        plt.ylabel('Зарплата, рублей', fontweight='bold', color='black', fontsize='12', horizontalalignment='center')
        plt.plot(x, y, '-o')
        # plt.show()
        plt.savefig(f'Средняя ЗП.png')
        img = openpyxl.drawing.image.Image(f'Средняя ЗП.png')
        self.insert_into_excel(query)
        wb = openpyxl.load_workbook(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")
        sheet = wb['Анализ']
        img = openpyxl.drawing.image.Image(f'Средняя ЗП.png')
        img.anchor = 'G1'
        sheet.add_image(img)
        wb.save(filename=f"{query.replace(' ', '_')}_vacancies_HH.xlsx")


if __name__ == '__main__':
    pass
